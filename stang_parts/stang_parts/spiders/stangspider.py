import scrapy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ..items import StangPartsItem  # Assuming this is your custom item class
import re

# Create a new Excel workbook and set up the worksheet
wb = Workbook()
ws = wb.active
ws.append(['title', 'price', 'references', 'availability'])

# Style the header row for better readability
fill = PatternFill(start_color='0070c1', end_color='0000CCFF', fill_type='solid')
ft = Font(bold=True, size=14, name='Comic Sans MS')
for row in ws["A1:D1"]:
    for cell in row:
        cell.fill = fill
        cell.font = ft

class StangPartsSpider(scrapy.Spider):
    """
    Web scraper for Stang Parts website to extract product information
    and save it to an Excel file.

    This spider crawls the Stang Parts website (https://stang-parts.de/en/)
    and extracts details like title, price, reference number, and availability
    for each car part listed on a product page. It then saves this data
    to an Excel file named "Supplies_for_a_car.xlsx".

    Attributes:
        name (str): Name of the spider for identification within Scrapy.
        allowed_domains (list): List of allowed domains for crawling to prevent
            accidental scraping of external sites.
        start_urls (list): List of starting URLs for the crawl.
    """

    name = "stang_parts"
    allowed_domains = ["stang-parts.de"]
    start_urls = [
        "https://stang-parts.de/en/2-strona-glowna",  # Starting URL for the crawl
    ]

    def save_to_excel(self, data):
        """
        Saves the extracted data to an Excel file.

        Args:
            data (dict): Dictionary containing the extracted product information.
        """

        lstsup = []
        for v in data.values():
            lstsup.append(v)
        ws.append(lstsup)
        wb.save("Supplies_for_a_car.xlsx")

    def parse(self, response):
        """
        Parses the product information from the given response object.

        Args:
            response (scrapy.http.Response): The response object from Scrapy
                containing the HTML content of the crawled page.

        Yields:
            dict: A dictionary containing the extracted product information
                for each car part on the page.
            scrapy.Request: Requests to follow links for further crawling.
        """

        for item in response.css('article.product-miniature.js-product-miniature'):
            # Extract product information from CSS selectors
            titles = item.css("h3.h3.product-title a::text").get()
            prices = item.css("span.price::text").get()
            references = item.css("p.pl_reference strong::text").get()
            availabilities = item.css("span.pl-availability::text")[1].get()

            # Clean availability text using regular expression
            availabilities = re.findall('^\s+(.+)\s+$', availabilities)[0]

            # Create a dictionary to store the extracted data
            data = {
                'titles': titles,
                'prices': prices,
                'references': references,
                'availabilities': availabilities,
            }

            # Yield the extracted data for later processing (e.g., saving)
            yield data

            # Follow pagination links and other relevant links for further crawling
            for href in response.css("ul.clearfix a::attr(href)"):
                yield response.follow(href, callback=self.parse)

        # Save data after processing all products on the current page
        self.save_to_excel(data)
