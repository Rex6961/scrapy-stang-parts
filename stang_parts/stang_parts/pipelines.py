from itemadapter import ItemAdapter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create a new Excel workbook and set up the worksheet
wb = Workbook()
ws = wb.active
ws.append(['title', 'price', 'references', 'availability'])

# Style the header row for better readability
fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type='solid')
ft = Font(bold=True, size=14, name='Comic Sans MS')

for row in ws["A1:D1"]:
    for cell in row:
        cell.fill = fill
        cell.font = ft

class StangPartsPipeline:
    seen_items = set()

    def process_item(self, item, spider):
        # Convert item values to a tuple for efficient hashing and comparison
        item_tuple = tuple(item.values())

        # Check if this item has been processed before
        if item_tuple not in self.seen_items:
            self.seen_items.add(item_tuple)
            ws.append(list(item_tuple))  # Append item values as a list
            wb.save("Supplies_for_a_car.xlsx")

        return item 